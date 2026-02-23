"""
pipeline.py — Shared pipeline functions for EPS Value Factor computation.

Mirrors the five-stage pipeline from the WifOR transitionvaluation project:
  Stage 1  Configuration       config.get_indicator_config()
  Stage 2  Data Loading        load_eps_sheet()
  Stage 3  Coefficient Matrix  create_coefficient_dataframe() + populate_coefficients()
  Stage 4  Inflation Adjust    calculate_inflation_factors() + apply_deflation()
  Stage 5  Output Export       save_results()

Derivation logic (re-implementing Excel formulas in Python)
──────────────────────────────────────────────────────────
The EPS formula for a substance s is:

  EPS_index[s] = Σ_{indicator i, pathway p} damage_cost[s, i, p]

where:

  damage_cost[s, i, p] = extent_of_impact[s, i, p]
                        × contribution_mean[s, i]      (= 1 / reference_stock)
                        × monetary_value[i]             (EUR / state-indicator-unit)

The Excel contains one row per (substance, indicator, pathway).  This module
reads all such rows, sums damage costs per substance, and uses that sum as
D[s] — the base damage-cost coefficient.

The WifOR formula then applies:

  C[y, s, c, n] = Sign × D[s] × I[y]

where I[y] = deflator[EU, y] / deflator[EU, base_year]  (EPS base year = 2015).

Because EPS 2015 is a global characterisation factor (D does not vary by
country), D[s] is broadcast uniformly across all 188 GeoRegion × NACE pairs.
"""

from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

from config import (
    EU_DEFLATOR_2015BASE,
    EU_DEFLATOR_BASE_YEAR,
    EU_DEFLATOR_LAST_KNOWN_YEAR,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# Stage 2 — Data Loading (special-case parsers)
# ═══════════════════════════════════════════════════════════════════════════════

def _load_fossil_resources(raw_rows: list[tuple]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Sheet 3: Fossil resources.

    Non-standard layout: computation worksheet, not a lookup table.
    Substance values are at fixed row/column positions derived from the Excel.

    Pathway columns (all with unit ELU/kg, base year 2015):
      Substance          row  col  description
      Fossil oil          14   4   damage cost EUR/kg
      Fossil coal         43   4   external + internal cost SUM (current technology)
      Lignite             46   4   damage cost EUR/kg
      Natural gas         59   4   total cost EUR/kg CH4
    """
    ENTRIES = [
        ("Fossil oil",   14, 4),
        ("Fossil coal",  43, 4),
        ("Lignite",      46, 4),
        ("Natural gas",  59, 4),
    ]
    pathway_records = []
    substance_records = []

    for substance, row_idx, col_idx in ENTRIES:
        if row_idx < len(raw_rows) and col_idx < len(raw_rows[row_idx]):
            val = raw_rows[row_idx][col_idx]
            if isinstance(val, (int, float)):
                pathway_records.append({
                    "substance": substance, "indicator": "total damage cost",
                    "unit": "kg", "pathway": "all",
                    "damage_cost_eur": float(val), "eps_index_source": float(val),
                })
                substance_records.append({
                    "substance": substance,
                    "eps_index_derived": float(val),
                    "eps_index_source": float(val),
                    "eps_index": float(val),
                })

    return pd.DataFrame(pathway_records), pd.DataFrame(substance_records).set_index("substance")


def _load_waste(raw_rows: list[tuple]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Sheet 16: Waste.

    Two substance rows at fixed positions:
      row 6  col 9  "Litter to ground"         (per number of items)
      row 7  col 9  "Plastic litter to water"  (per kg)
    """
    ENTRIES = [
        (6, "Litter to ground"),
        (7, "Plastic litter to water"),
    ]
    pathway_records = []
    substance_records = []

    for row_idx, substance in ENTRIES:
        if row_idx < len(raw_rows) and len(raw_rows[row_idx]) > 9:
            val = raw_rows[row_idx][9]
            if isinstance(val, (int, float)):
                pathway_records.append({
                    "substance": substance, "indicator": "total damage cost",
                    "unit": "unit", "pathway": "all",
                    "damage_cost_eur": float(val), "eps_index_source": float(val),
                })
                substance_records.append({
                    "substance": substance,
                    "eps_index_derived": float(val),
                    "eps_index_source": float(val),
                    "eps_index": float(val),
                })

    return pd.DataFrame(pathway_records), pd.DataFrame(substance_records).set_index("substance")


def _load_radionuclides(raw_rows: list[tuple]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Sheet 14: Radionuclides.

    Structure:
      Row 0  section title
      Row 1  header: [..., 'Collective dose per unit release', ..., 'Damage cost, EUR/TBq', ...]
      Row 2  unit row (manSv/TBq …)
      Row 3+ data rows: col 0 = nuclide name, col 6 = damage cost (EUR/TBq)

    Stops at the 'Emissions to water' section title.
    Water emissions are marked 'negligable impact' and are excluded.
    """
    DAMAGE_COST_COL = 6
    pathway_records = []
    substance_records = []

    for row in raw_rows[3:]:  # skip title, header, unit rows
        if not any(v is not None for v in row):
            continue
        name = _clean(row[0])
        if not name:
            continue
        # Stop at the water-emissions section header
        if "emission" in name.lower() and len(name) > 10:
            break
        # Skip negligible-impact rows (no damage cost value)
        if DAMAGE_COST_COL >= len(row) or not isinstance(row[DAMAGE_COST_COL], (int, float)):
            continue
        val = float(row[DAMAGE_COST_COL])
        pathway_records.append({
            "substance": name, "indicator": "YOLL + cancer",
            "unit": "TBq", "pathway": "all",
            "damage_cost_eur": val, "eps_index_source": val,
        })
        substance_records.append({
            "substance": name,
            "eps_index_derived": val,
            "eps_index_source": val,
            "eps_index": val,
        })

    pathway_df = pd.DataFrame(pathway_records)
    substance_df = (
        pd.DataFrame(substance_records).set_index("substance")
        if substance_records else pd.DataFrame()
    )
    return pathway_df, substance_df


# Map sheet names → special parser functions (called instead of the generic parser)
_SPECIAL_SHEET_LOADERS = {
    "3. Fossil res":    _load_fossil_resources,
    "14. Radionuclids": _load_radionuclides,
    "16. waste":        _load_waste,
}


# ═══════════════════════════════════════════════════════════════════════════════
# Stage 2 — Data Loading (generic parser)
# ═══════════════════════════════════════════════════════════════════════════════

def _clean(value: Any) -> str:
    """Normalise a cell value to a stripped, single-space string."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _is_header_row(row: tuple) -> bool:
    """True if the first non-None cell looks like a column header."""
    first = _clean(row[0]).lower() if row else ""
    keywords = [
        "substance", "emission", "element", "flow group",
        "land use", "radionuclid", "resource", "indicator",
    ]
    return bool(first) and any(kw in first for kw in keywords)


def _find_col(header: tuple, *keywords: str) -> int:
    """
    Return the first column index whose header cell contains any of the keywords.
    Two-pass: high-priority keywords first, then fallback.
    """
    HIGH = ["eps", "default index", "elu/", "weighting factor"]
    FALLBACK = ["damage cost"]
    for kw_set in (HIGH, FALLBACK):
        for i, cell in enumerate(header):
            low = _clean(cell).lower()
            if any(k in low for k in kw_set):
                return i
    return -1


def load_eps_sheet(
    xlsx_path: str | Path,
    sheet_name: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Read one EPS 2015 worksheet and return two DataFrames:

    ``pathway_df``
        Full pathway-level data (one row per substance × indicator × pathway).
        Columns: substance, indicator, unit, pathway,
                 extent_mean, extent_unc, contribution_mean, contribution_unc,
                 pathway_cf, cf_state_indicator, damage_cost_eur, eps_index.

    ``substance_df``
        Substance-level EPS index, derived by summing damage_cost_eur
        across all (indicator, pathway) rows per substance.
        Index: substance name.
        Columns: eps_index_derived (re-derived from pathway sums),
                 eps_index_source  (as given in the "all / all" summary row).

    The re-derived value validates against the source value; any discrepancy
    > 0.1 % is logged as a warning.
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path.name}")

    ws = wb[sheet_name]
    raw_rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
    wb.close()

    # ── Dispatch to special parsers for non-standard sheet layouts ──
    if sheet_name in _SPECIAL_SHEET_LOADERS:
        pathway_df, substance_df = _SPECIAL_SHEET_LOADERS[sheet_name](raw_rows)
        logger.info(
            "  Loaded sheet %-28s → %3d substances (special parser)",
            f"'{sheet_name}'", len(substance_df),
        )
        return pathway_df, substance_df

    # ── Generic parser: locate header row ──
    header_idx = next(
        (i for i, r in enumerate(raw_rows) if r and _is_header_row(r)), -1
    )
    if header_idx < 0:
        logger.warning("No header found in sheet '%s' — returning empty DataFrames", sheet_name)
        return pd.DataFrame(), pd.DataFrame()

    header = raw_rows[header_idx]
    eps_col = _find_col(header)

    # ── Build column name → index map from header ──
    def _hcol(keywords: list[str]) -> int:
        for i, cell in enumerate(header):
            low = _clean(cell).lower()
            if any(k in low for k in keywords):
                return i
        return -1

    COL = {
        "substance":         0,
        "indicator":         _hcol(["indicator", "indicators", "element name", "substance name"]),
        "unit":              _hcol(["unit"]),
        "pathway":           _hcol(["pathway"]),
        "extent_mean":       _hcol(["extent of impact", "extent", "ld50", "collective"]),
        "extent_unc":        _hcol(["extent of impact\n un", "uncertainty\n factor"]),
        "contribution_mean": _hcol(["contribution mean"]),
        "contribution_unc":  _hcol(["contribution\n unc"]),
        "pathway_cf":        _hcol(["pathway specific", "characterisation\n factor for state"]),
        "damage_cost":       _hcol(["damage cost"]),
        "eps_index":         eps_col,
    }

    # ── Walk data rows ──
    records = []
    current_substance: str = ""
    SUMMARY_KWS = {"all", "total", "sum"}

    for row in raw_rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue

        col0 = _clean(row[0])

        # Carry substance name forward (merged-cell pattern)
        if col0 and col0.lower() not in SUMMARY_KWS:
            current_substance = col0

        if not current_substance:
            continue

        def _get(col: int):
            return row[col] if col >= 0 and col < len(row) else None

        records.append({
            "substance":         current_substance,
            "indicator":         _get(COL["indicator"]),
            "unit":              _get(COL["unit"]),
            "pathway":           _get(COL["pathway"]),
            "extent_mean":       _get(COL["extent_mean"]),
            "extent_unc":        _get(COL["extent_unc"]),
            "contribution_mean": _get(COL["contribution_mean"]),
            "contribution_unc":  _get(COL["contribution_unc"]),
            "pathway_cf":        _get(COL["pathway_cf"]),
            "damage_cost_eur":   _get(COL["damage_cost"]),
            "eps_index_source":  _get(COL["eps_index"]),
        })

    if not records:
        return pd.DataFrame(), pd.DataFrame()

    pathway_df = pd.DataFrame(records)

    # ── Derive EPS index from pathway sums ──
    # The EPS index = sum of damage_cost_eur across all non-summary rows
    is_summary = pathway_df["indicator"].apply(
        lambda v: str(v).lower().strip() in SUMMARY_KWS if v is not None else False
    )
    pathway_rows = pathway_df[~is_summary].copy()
    pathway_rows["damage_cost_eur"] = pd.to_numeric(
        pathway_rows["damage_cost_eur"], errors="coerce"
    ).fillna(0)

    derived = (
        pathway_rows.groupby("substance")["damage_cost_eur"]
        .sum()
        .rename("eps_index_derived")
    )

    # ── Extract source EPS index from the "all/all" summary rows ──
    summary_rows = pathway_df[is_summary & pathway_df["eps_index_source"].notna()]
    source = (
        summary_rows.drop_duplicates("substance")
        .set_index("substance")["eps_index_source"]
    )

    substance_df = pd.DataFrame({
        "eps_index_derived": derived,
        "eps_index_source":  source,
    })

    # ── Validate: re-derived vs source ──
    both = substance_df.dropna()
    if not both.empty:
        rel_err = ((both["eps_index_derived"] - both["eps_index_source"])
                   / both["eps_index_source"].replace(0, np.nan)).abs()
        bad = rel_err[rel_err > 0.001]
        if not bad.empty:
            logger.warning(
                "Sheet '%s': %d substance(s) have >0.1%% derivation error: %s",
                sheet_name, len(bad), bad.index.tolist(),
            )

    # Use derived value where available, fall back to source
    substance_df["eps_index"] = substance_df["eps_index_derived"].combine_first(
        substance_df["eps_index_source"]
    )

    logger.info(
        "  Loaded sheet %-28s → %3d substances (%d with full pathway data)",
        f"'{sheet_name}'", len(substance_df), len(derived),
    )
    return pathway_df, substance_df


# ═══════════════════════════════════════════════════════════════════════════════
# Stage 3 — Coefficient Matrix
# ═══════════════════════════════════════════════════════════════════════════════

def _make_variable_name(category: str, substance: str, unit: str) -> str:
    """
    Build a WifOR-style variable name.

    WifOR example: "COEFFICIENT AirEmission_NH3, in USD (WifOR)"
    EPS  example:  "EPS_InorganicGases_CO2, in ELU/kg (Steen2015)"
    """
    cat_slug = category.replace(" ", "").replace("-", "")
    sub_slug = substance.replace(" ", "_").replace(",", "").replace("/", "per")
    return f"EPS_{cat_slug}_{sub_slug}, in {unit} (Steen2015)"


def create_coefficient_dataframe(
    years: list[str],
    variables: list[str],
    countries: list[str],
    nace_sectors: list[str],
    initial_value: float = 1.0,
) -> pd.DataFrame:
    """
    Create the empty coefficient matrix C[y, variable, country, sector].

    Shape: (len(years) × len(variables))  ×  (len(countries) × len(nace_sectors))

    Row MultiIndex:    (Year, Variable)    — names match WifOR convention
    Column MultiIndex: (GeoRegion, NACE)   — names match WifOR convention

    All cells initialised to ``initial_value`` (typically the sign: ±1.0).

    Mirrors WifOR's ``create_coefficient_dataframe()``.
    """
    row_idx = pd.MultiIndex.from_product(
        [years, variables], names=["Year", "Variable"]
    )
    col_idx = pd.MultiIndex.from_product(
        [countries, nace_sectors], names=["GeoRegion", "NACE"]
    )
    return pd.DataFrame(
        initial_value,
        index=row_idx,
        columns=col_idx,
        dtype=float,
    )


def populate_coefficients(
    coeff: pd.DataFrame,
    substance_df: pd.DataFrame,
    years: list[str],
    variables: list[str],
    sign: float,
) -> pd.DataFrame:
    """
    Fill in D[substance] values, broadcast uniformly across all (GeoRegion, NACE).

    Uses vectorised numpy operations: for each year, assigns all substance
    rows at once via a 2-D slice instead of looping over individual (year, var)
    cells.

    Formula applied here (before inflation):
        C[y, substance, c, n] = sign × EPS_index[substance]
    """
    d_values = sign * substance_df["eps_index"].values   # (N_sub,)
    N_sub = len(d_values)
    N_col = coeff.shape[1]
    arr = coeff.to_numpy(copy=False)                     # numpy view (no copy)

    for i_year in range(len(years)):
        row_start = i_year * N_sub
        row_end   = row_start + N_sub
        # Broadcast d_values across all country×sector columns
        arr[row_start:row_end, :] = d_values[:, np.newaxis]

    return coeff


# ═══════════════════════════════════════════════════════════════════════════════
# Stage 4 — Inflation Adjustment
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_inflation_factors(
    years: list[str],
    base_year: str = EU_DEFLATOR_BASE_YEAR,
) -> pd.Series:
    """
    Compute year-specific inflation factors I[y] for the EU deflator.

    I[y] = deflator[EU, y] / deflator[EU, base_year]

    For forecast years beyond the last known deflator year, the factor
    is frozen at the last known value (mirrors WifOR behaviour).

    Returns
    -------
    pd.Series
        Index = year strings; values = normalised inflation factors (base_year = 1.0).
    """
    factors: dict[str, float] = {}
    last_known_val = EU_DEFLATOR_2015BASE[EU_DEFLATOR_LAST_KNOWN_YEAR]
    base_val = EU_DEFLATOR_2015BASE[base_year]

    for year in years:
        raw = EU_DEFLATOR_2015BASE.get(year, last_known_val)
        factors[year] = raw / base_val

    series = pd.Series(factors, name="inflation_factor")
    logger.debug("Inflation factors (EU deflator, base %s): %s", base_year, series.to_dict())
    return series


def apply_deflation(
    coeff: pd.DataFrame,
    inflation_factors: pd.Series,
    years: list[str],
    variables: list[str],
) -> pd.DataFrame:
    """
    Apply I[y] to each year slice of the coefficient matrix.

    After this step:
        C[y, substance, c, n] = sign × EPS_index[substance] × I[y]

    Vectorised: builds a column-vector of per-row inflation factors and
    multiplies the entire numpy array in one broadcast operation.

    Mirrors WifOR's ``apply_deflation()``.
    """
    coeff_final = coeff.copy()
    arr = coeff_final.to_numpy(copy=False)
    N_sub = len(variables)

    # Build inflation multiplier vector: shape (N_years * N_sub, 1)
    i_col = np.repeat(
        [inflation_factors[y] for y in years],
        N_sub,
    ).reshape(-1, 1)

    arr *= i_col          # broadcast: (N_rows, N_cols) *= (N_rows, 1)
    return coeff_final


def build_unit_frame(
    variables: list[str],
    years: list[str],
    unit_base: str,
) -> pd.DataFrame:
    """
    Build the units metadata DataFrame.

    For years with real deflator data  → "2015EUR/{unit_base}"  (base year nominal)
    For forecast years                 → "2023EUR/{unit_base}"  (frozen at last known)

    Mirrors WifOR's ``units`` DataFrame (index=variables, columns=years).
    """
    real_years = set(EU_DEFLATOR_2015BASE)
    last_known = EU_DEFLATOR_LAST_KNOWN_YEAR
    # Unit string format: "{year}ELU/{physical_unit}"
    # e.g. unit_base="ELU/kg" → strip leading "ELU/" to get "kg" → "2015ELU/kg"
    phys_unit = unit_base.lstrip("ELU/") if unit_base.startswith("ELU/") else unit_base
    rows = {}
    for var in variables:
        rows[var] = {
            y: f"{y}ELU/{phys_unit}" if y in real_years else f"{last_known}ELU/{phys_unit}"
            for y in years
        }
    return pd.DataFrame(rows).T


# ═══════════════════════════════════════════════════════════════════════════════
# Stage 5 — Output Export
# ═══════════════════════════════════════════════════════════════════════════════

def save_results(
    coeff_final: pd.DataFrame,
    units: pd.DataFrame,
    pathway_df: pd.DataFrame,
    hdf5_path: str | Path,
    excel_path: str | Path,
    hdf5_coeff_key: str = "coefficient",
    hdf5_unit_key: str = "unit",
    excel_sheet_coeff: str = "Coefficients",
    excel_sheet_units: str = "Units",
    excel_sheet_pathways: str = "Pathway data",
    freeze_coeff: tuple = (1, 2),
    freeze_units: tuple = (1, 1),
    excel_max_cols: int = 50,
) -> None:
    """
    Persist coefficient matrix and unit metadata to HDF5 and Excel.

    HDF5 keys (mirrors WifOR):
      "coefficient"  → full coeff_final DataFrame (all GeoRegion × NACE columns)
      "unit"         → units DataFrame

    Excel sheets:
      "Coefficients"   → compact view: first ``excel_max_cols`` country columns.
                         Because EPS coefficients are globally uniform (D[s] does
                         not vary by country), all columns hold the same value —
                         showing a representative subset saves file size and write
                         time without losing information.
      "Units"          → units metadata
      "Pathway data"   → full Excel derivation logic from source XLSX

    Parameters mirror WifOR's ``save_results()`` signature.
    """
    hdf5_path = Path(hdf5_path)
    excel_path = Path(excel_path)
    hdf5_path.parent.mkdir(parents=True, exist_ok=True)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # ── HDF5 — full matrix ──────────────────────────────────────────────────
    coeff_final.to_hdf(hdf5_path, key=hdf5_coeff_key, mode="w", complevel=4, complib="blosc")
    units.to_hdf(hdf5_path, key=hdf5_unit_key, mode="a")
    logger.info("Saved HDF5: %s", hdf5_path)

    # ── Excel — compact representative view ─────────────────────────────────
    # Trim columns: one value per NACE sector for the first country (representative)
    # since all countries hold the same coefficient in EPS 2015.
    n_cols = coeff_final.shape[1]
    if n_cols > excel_max_cols:
        coeff_excel = coeff_final.iloc[:, :excel_max_cols]
        col_note = (
            f"Note: Showing {excel_max_cols} of {n_cols} GeoRegion×NACE columns. "
            f"All columns hold the same value (EPS 2015 is a global CF — "
            f"no country variation). Full data in the companion .h5 file."
        )
    else:
        coeff_excel = coeff_final
        col_note = None

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        coeff_excel.to_excel(
            writer,
            sheet_name=excel_sheet_coeff,
            merge_cells=False,
            freeze_panes=freeze_coeff,
        )
        # Add column-truncation note in cell A1 of a Notes sheet
        if col_note:
            note_df = pd.DataFrame({"Note": [col_note]})
            note_df.to_excel(writer, sheet_name="Notes", index=False)

        units.to_excel(
            writer,
            sheet_name=excel_sheet_units,
            merge_cells=False,
            freeze_panes=freeze_units,
        )
        if not pathway_df.empty:
            pathway_df.to_excel(
                writer,
                sheet_name=excel_sheet_pathways,
                index=False,
            )
    logger.info("Saved Excel: %s", excel_path)


# ═══════════════════════════════════════════════════════════════════════════════
# Convenience: full indicator run
# ═══════════════════════════════════════════════════════════════════════════════

def run_indicator(indicator_key: str, xlsx_path: str | Path | None = None) -> dict:
    """
    Run the full 5-stage pipeline for a single EPS indicator (sheet).

    This is the function called by each individual indicator script.
    It mirrors the body of a WifOR indicator script.

    Parameters
    ----------
    indicator_key:
        Key into config.INDICATORS, e.g. ``"inorganic_gases"``.
    xlsx_path:
        Override path to the EPS XLSX (defaults to config value).

    Returns
    -------
    dict with keys: coeff_final, units, pathway_df, substance_df, variables
    """
    import config  # local import so indicator scripts don't need sys.path tricks

    # Stage 1 — Configuration
    cfg = config.get_indicator_config(indicator_key)
    years = cfg["years"]
    countries = cfg["countries"]
    nace = cfg["nace"]
    sign = cfg["coefficient_sign"]
    unit_base = cfg["unit_base"]
    sheet_name = cfg["sheet_name"]
    category_label = indicator_key.replace("_", " ").title()

    if xlsx_path is None:
        xlsx_path = cfg["eps_xlsx_with_secondary"]

    logger.info("=" * 60)
    logger.info("Indicator : %s", indicator_key)
    logger.info("Sheet     : %s", sheet_name)
    logger.info("Sign      : %+.1f", sign)
    logger.info("Years     : %s … %s (%d)", years[0], years[-1], len(years))
    logger.info("Countries : %d | Sectors: %d", len(countries), len(nace))

    # Stage 2 — Data Loading
    pathway_df, substance_df = load_eps_sheet(xlsx_path, sheet_name)

    if substance_df.empty or "eps_index" not in substance_df.columns:
        logger.error("No EPS data extracted for indicator '%s'", indicator_key)
        return {}

    substance_df = substance_df.dropna(subset=["eps_index"])
    substances = substance_df.index.tolist()

    # Build WifOR-style variable names
    variables = [
        _make_variable_name(category_label, s, unit_base)
        for s in substances
    ]

    logger.info("Substances: %d extracted", len(substances))

    # Stage 3 — Coefficient Matrix
    coeff = create_coefficient_dataframe(years, variables, countries, nace, initial_value=1.0)
    coeff = populate_coefficients(coeff, substance_df, years, variables, sign)

    # Stage 4 — Inflation Adjustment
    inflation_factors = calculate_inflation_factors(years)
    coeff_final = apply_deflation(coeff, inflation_factors, years, variables)
    units = build_unit_frame(variables, years, unit_base)

    # Stage 5 — Output Export
    save_results(
        coeff_final, units, pathway_df,
        hdf5_path=cfg["hdf5_path"],
        excel_path=cfg["excel_path"],
        hdf5_coeff_key=cfg["hdf5_coeff_key"],
        hdf5_unit_key=cfg["hdf5_unit_key"],
        excel_sheet_coeff=cfg["excel_sheet_coefficients"],
        excel_sheet_units=cfg["excel_sheet_units"],
        freeze_coeff=cfg["excel_freeze_coefficients"],
        freeze_units=cfg["excel_freeze_units"],
    )

    logger.info(
        "Done %-20s → C matrix shape %s",
        indicator_key, coeff_final.shape,
    )

    return {
        "coeff_final":   coeff_final,
        "units":         units,
        "pathway_df":    pathway_df,
        "substance_df":  substance_df,
        "variables":     variables,
    }
